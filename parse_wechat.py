#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
minquzi - 微信聊天记录业绩提取工具

本脚本接收 AI 识别好的结构化 JSON 数据，按规则换算字段并生成 Excel。

使用方式：
    python parse_wechat.py --from-json data.json --output output/

JSON 数据格式（由 AI 识别填充）：
[
  {
    "date": "4月30日",
    "group": "广深杭果兮",
    "tail": "6514",
    "raw_perf": "450.0",
    "director": "",
    "reward": "",
    "ai_abnormal": false,
    "raw_text": "整段识别文字..."
  },
  ...
]
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import yaml
except ImportError:
    print("缺少依赖: pip install pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).parent.resolve()
CONFIG_PATH = SCRIPT_DIR / "config.yaml"


def load_config(path: Path = CONFIG_PATH) -> dict:
    """加载 config.yaml"""
    if not path.exists():
        raise FileNotFoundError(f"配置文件不存在: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_group_multiplier(group_name: str, special_groups: list) -> int:
    """
    根据群组名称获取特殊倍率
    模糊匹配：群组名称包含 special_groups[i].name 即命中
    未命中返回 1
    """
    if not group_name:
        return 1
    for rule in special_groups or []:
        keyword = rule.get("name", "")
        if keyword and keyword in group_name:
            return int(rule.get("multiplier", 1))
    return 1


def get_group_point(group_name: str, group_points: dict):
    """
    根据群组名称获取点位（精确匹配）
    未配置返回 None
    """
    if not group_name or not group_points:
        return None
    return group_points.get(group_name)


def calc_actual_perf(raw_perf_str: str, group_name: str, special_groups: list):
    """
    根据原始业绩字符串计算实际业绩
    规则：
      - 有小数点：×10
      - 无小数点：原样
      - 特殊群组：再 × multiplier
    返回 int 或 None
    """
    if raw_perf_str is None or raw_perf_str == "":
        return None

    s = str(raw_perf_str).strip()
    try:
        if "." in s:
            base = float(s) * 10
        else:
            base = float(s)
    except ValueError:
        return None

    multiplier = get_group_multiplier(group_name, special_groups)
    actual = base * multiplier
    return int(round(actual))


def calc_settlement(actual_perf, point):
    """应结 = 实际业绩 × 点位"""
    if actual_perf is None or point is None:
        return None
    try:
        return round(float(actual_perf) * float(point), 2)
    except (TypeError, ValueError):
        return None


def clean_group_name(raw_title: str) -> str:
    """
    清洗群组名称：去掉末尾的 (数字) 或 （数字）
    示例："天津南开和平门店(126)" → "天津南开和平门店"
    """
    if not raw_title:
        return ""
    # 匹配末尾的 (数字) 或 （数字）
    cleaned = re.sub(r"\s*[（(]\s*\d+\s*[)）]\s*$", "", raw_title.strip())
    return cleaned.strip()


def process_records(records: list, config: dict) -> list:
    """
    处理记录列表，计算实际业绩、应结、AI识别异常等衍生字段
    """
    special_groups = config.get("special_groups", [])
    group_points = config.get("group_points", {})

    results = []
    for rec in records:
        date = rec.get("date", "")
        group = clean_group_name(rec.get("group", ""))
        tail = rec.get("tail", "")
        raw_perf = rec.get("raw_perf", "")
        director = rec.get("director", "")
        reward = rec.get("reward", "")
        ai_abnormal_flag = rec.get("ai_abnormal", False)
        raw_text = rec.get("raw_text", "")

        actual_perf = calc_actual_perf(raw_perf, group, special_groups)
        point = get_group_point(group, group_points)
        settlement = calc_settlement(actual_perf, point)

        # 异常判断：若未显式标记，按业绩是否能解析判断
        if ai_abnormal_flag:
            ai_abnormal = "是"
        elif raw_perf == "" and raw_text:
            # 有原始文字但没业绩 → 异常
            ai_abnormal = "是"
        else:
            ai_abnormal = "否"

        results.append({
            "日期": date,
            "群组": group,
            "尾号": tail,
            "原始业绩": raw_perf if raw_perf != "" else "",
            "实际业绩": actual_perf if actual_perf is not None else "",
            "点位": point if point is not None else "",
            "应结": settlement if settlement is not None else "",
            "总监": director,
            "奖励": reward,
            "AI识别异常": ai_abnormal,
            "原始数据": raw_text,
        })
    return results


def write_excel(rows: list, config: dict, output_dir: Path, date_str: str = None) -> Path:
    """
    将处理好的数据写入 Excel
    """
    output_cfg = config.get("output", {}) or {}
    prefix = output_cfg.get("filename_prefix", "业绩汇总")
    columns = output_cfg.get("columns") or [
        "日期", "群组", "尾号", "原始业绩", "实际业绩",
        "点位", "应结", "总监", "奖励", "AI识别异常", "原始数据"
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_date = (date_str or "").replace("/", "-").replace("\\", "-") or "nodate"
    filename = f"{prefix}_{safe_date}_{ts}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "业绩汇总"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    abnormal_fill = PatternFill("solid", fgColor="FFD966")  # 黄色高亮异常行
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 写表头
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 写数据
    for r_idx, row in enumerate(rows, start=2):
        is_abnormal = row.get("AI识别异常") == "是"
        for c_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if col == "原始数据":
                cell.alignment = left_wrap
            else:
                cell.alignment = center
            if is_abnormal:
                cell.fill = abnormal_fill

    # 列宽
    width_map = {
        "日期": 10, "群组": 24, "尾号": 8, "原始业绩": 10, "实际业绩": 10,
        "点位": 8, "应结": 10, "总监": 10, "奖励": 10,
        "AI识别异常": 12, "原始数据": 60,
    }
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_map.get(col, 15)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


def main():
    parser = argparse.ArgumentParser(description="minquzi - 微信业绩提取生成 Excel")
    parser.add_argument("--from-json", required=True, help="AI 识别好的 JSON 文件路径")
    parser.add_argument("--output", default="output", help="输出目录（默认 output/）")
    parser.add_argument("--date", default="", help="日期标签（用于文件名）")
    args = parser.parse_args()

    config = load_config()

    json_path = Path(args.from_json)
    if not json_path.exists():
        print(f"[错误] JSON 文件不存在: {json_path}", file=sys.stderr)
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    if not isinstance(records, list):
        print("[错误] JSON 根节点必须是数组", file=sys.stderr)
        sys.exit(1)

    rows = process_records(records, config)

    output_dir = Path(args.output)
    if not output_dir.is_absolute():
        output_dir = SCRIPT_DIR / output_dir

    filepath = write_excel(rows, config, output_dir, date_str=args.date)
    print(f"[成功] 已生成 Excel: {filepath}")
    print(f"[统计] 共 {len(rows)} 条记录，其中异常 {sum(1 for r in rows if r['AI识别异常']=='是')} 条")


if __name__ == "__main__":
    main()
